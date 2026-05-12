"""
Extract parts data + embedded images from InventoryFlow_Data.xlsx.

Outputs:
  images/                — every embedded image (PNG/JPG) from the workbook
  parts.db               — SQLite DB with one row per part_number
  summary.json           — totals + sheet list
  sample_data.json       — first 100 parts (pretty-printed for Notepad)
"""
import sys, io, os, json, sqlite3, re, zipfile, shutil
import xml.etree.ElementTree as ET
from collections import defaultdict, OrderedDict
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

XLSX = "InventoryFlow_Data.xlsx"
IMG_DIR = "images"
DB_PATH = "parts.db"
SUMMARY_PATH = "summary.json"
SAMPLE_PATH = "sample_data.json"
SAMPLE_LIMIT = 100

# Sheets that aren't parts catalogs — we'll still record their names in summary,
# but we won't try to extract part rows from them.
NON_PART_SHEETS = {
    "TABLE OF CONTENTS", "Sheet18",
    "Fork seal specs ", "ATV Wheel specs", "Spoke Specs",
    "SPARK PLUGS", "Battery specs ",
    "Owners manuals ", "owners manuals",
    "DirtbikePitbike wheel bearingse",
    "SNOW TRACK KIT",
    "Carburetor Jets",
    "ski kit parts",
    "eA110 upgrade kit ",
}

# Priority-ranked header matchers. Higher rank wins when multiple columns match.
# (substring or exact match, rank). Rank counts down — lower number = higher priority.
PN_RULES = [           # part number column
    ("new part number", 1),
    ("新编码", 1),
    ("part number", 2),
    ("u8 code", 3),
    ("u8编码", 3),
    ("factory's part no", 4),
    ("parts no.", 5),
    ("parts no", 5),
    ("factory code", 6),
    ("tx warehouse no", 7),
]
EN_RULES = [           # english description column
    ("en name", 1),
    ("english description", 2),
    ("description", 3),
]
CN_RULES = [           # chinese description column
    ("cn name", 1),
    ("chinese description", 2),
    ("specifications in cn", 3),
    ("中文名称", 3),
    ("物料名称", 3),
]

def _match_rule(value, rules):
    """Return (rank, matched) for value; lower rank = better. None if no match."""
    for needle, rank in rules:
        if needle in value:
            return rank
    return None

def find_header(row):
    """Return dict of {field -> col_index} if the row looks like a header, else None."""
    if row is None:
        return None
    norm = [(str(c).strip().lower().replace("\n", " ") if c is not None else "") for c in row]
    best = {"pn": None, "en": None, "cn": None, "price": None}
    best_rank = {"pn": 99, "en": 99, "cn": 99, "price": 99}

    for i, v in enumerate(norm):
        if not v:
            continue
        for field, rules in (("pn", PN_RULES), ("en", EN_RULES), ("cn", CN_RULES)):
            r = _match_rule(v, rules)
            if r is not None and r < best_rank[field]:
                best_rank[field] = r
                best[field] = i
        if "retail" in v and best["price"] is None:
            best["price"] = i

    if best["pn"] is None or best["en"] is None:
        return None
    return {k: v for k, v in best.items() if v is not None}


YEAR_PATTERNS = [
    re.compile(r"\((\d{4})\s*-\s*(\d{4})\)"),    # (2016-2020)
    re.compile(r"\((\d{4})\+\)"),                # (2021+)
    re.compile(r"(\d{4})\+"),                    # 2021+
    re.compile(r"\((\d{4})\)"),                  # (2024)
    re.compile(r"\b(20\d{2})\b"),                # bare 2024
]

def derive_fitment(sheet_name):
    """Return one fitment dict for this sheet — {year, make, model}."""
    s = sheet_name.strip()
    years = None
    for pat in YEAR_PATTERNS:
        m = pat.search(s)
        if not m:
            continue
        groups = [g for g in m.groups() if g]
        if len(groups) == 2:
            years = f"{groups[0]}-{groups[1]}"
        else:
            years = groups[0] + ("+" if "+" in m.group(0) else "")
        break
    # Strip year fragments to clean up the model label
    model = s
    for pat in YEAR_PATTERNS:
        model = pat.sub("", model)
    model = re.sub(r"\s+", " ", model).strip(" -")
    # Make is not labeled in the workbook — leave null
    return {"year": years, "make": None, "model": model, "sheet": sheet_name.strip()}


# ----------------------------------------------------------------------------
# Step 1: extract all media files and build sheet → [(row_idx, image_file)] map
# ----------------------------------------------------------------------------
NS_REL = "{http://schemas.openxmlformats.org/package/2006/relationships}"
NS_DR = "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}"
NS_R = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
NS_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

def extract_images_and_anchors(xlsx_path):
    if os.path.isdir(IMG_DIR):
        shutil.rmtree(IMG_DIR)
    os.makedirs(IMG_DIR, exist_ok=True)

    with zipfile.ZipFile(xlsx_path) as z:
        names = z.namelist()

        # Dump every media file
        media = [n for n in names if n.startswith("xl/media/")]
        for m in media:
            data = z.read(m)
            out = os.path.join(IMG_DIR, os.path.basename(m))
            with open(out, "wb") as f:
                f.write(data)

        # workbook.xml — sheet name → r:id
        wb_xml = ET.fromstring(z.read("xl/workbook.xml"))
        sheet_rid = {}
        for s in wb_xml.iter(NS_MAIN + "sheet"):
            sheet_rid[s.attrib["name"]] = s.attrib[NS_R + "id"]

        # workbook rels — r:id → sheet target
        wb_rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rid_target = {}
        for r in wb_rels.iter(NS_REL + "Relationship"):
            rid_target[r.attrib["Id"]] = r.attrib["Target"]

        # For each sheet, find its drawing file via worksheet rels
        sheet_drawing = {}
        for sname, rid in sheet_rid.items():
            target = rid_target.get(rid, "")
            sheet_path = os.path.normpath(os.path.join("xl", target)).replace("\\", "/")
            rels_path = sheet_path.replace("worksheets/", "worksheets/_rels/") + ".rels"
            if rels_path not in names:
                continue
            sheet_rels = ET.fromstring(z.read(rels_path))
            for rel in sheet_rels.iter(NS_REL + "Relationship"):
                if rel.attrib.get("Type", "").endswith("/drawing"):
                    drawing_target = rel.attrib["Target"]
                    drawing_path = os.path.normpath(
                        os.path.join(os.path.dirname(sheet_path), drawing_target)
                    ).replace("\\", "/")
                    sheet_drawing[sname] = drawing_path

        # For each drawing, parse anchors → (row, image_filename)
        sheet_anchors = defaultdict(list)
        for sname, drawing_path in sheet_drawing.items():
            if drawing_path not in names:
                continue
            drawing_xml = ET.fromstring(z.read(drawing_path))
            drels_path = drawing_path.replace("drawings/", "drawings/_rels/") + ".rels"
            rid_image = {}
            if drels_path in names:
                drels = ET.fromstring(z.read(drels_path))
                for rel in drels.iter(NS_REL + "Relationship"):
                    if "/image" in rel.attrib.get("Type", ""):
                        rid_image[rel.attrib["Id"]] = os.path.basename(rel.attrib["Target"])
            for anchor in list(drawing_xml):
                if not anchor.tag.endswith("Anchor"):
                    continue
                frm = anchor.find(NS_DR + "from")
                if frm is None:
                    continue
                row_el = frm.find(NS_DR + "row")
                if row_el is None:
                    continue
                from_row = int(row_el.text)
                blip = anchor.find(".//" + "{http://schemas.openxmlformats.org/drawingml/2006/main}blip")
                if blip is None:
                    continue
                rid = blip.attrib.get(NS_R + "embed") or blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
                if rid and rid in rid_image:
                    sheet_anchors[sname].append((from_row, rid_image[rid]))
            sheet_anchors[sname].sort()
    return sheet_anchors


# ----------------------------------------------------------------------------
# Step 2: walk every sheet, pull part rows, build the merged record table.
# ----------------------------------------------------------------------------
def first_non_none(*vals):
    for v in vals:
        if v is not None and (not isinstance(v, str) or v.strip() != ""):
            return v
    return None

def normalize_pn(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

def to_price(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def extract_parts(xlsx_path, sheet_anchors):
    print(f"Loading workbook (read_only) ...")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    parts = OrderedDict()        # part_number -> dict
    sheet_part_counts = {}
    sheet_names = list(wb.sheetnames)

    for sname in sheet_names:
        if sname in NON_PART_SHEETS:
            sheet_part_counts[sname] = 0
            continue
        ws = wb[sname]
        anchors = sheet_anchors.get(sname, [])
        # Sorted list of (row_idx_0based, image_filename)
        anchor_rows = [a[0] for a in anchors]
        anchor_imgs = [a[1] for a in anchors]

        fitment_entry = derive_fitment(sname)
        current_header = None
        added = 0

        for row_idx_0, row in enumerate(ws.iter_rows(values_only=True)):
            # Re-detect header rows (sheets have multiple sub-tables).
            hdr = find_header(row)
            if hdr:
                current_header = hdr
                continue
            if current_header is None:
                continue

            pn = normalize_pn(row[current_header["pn"]] if current_header["pn"] < len(row) else None)
            if not pn:
                continue
            # Skip rows that are obviously not part numbers (pure digits like "1", "2")
            if pn.isdigit() and len(pn) <= 3:
                continue

            en = first_non_none(
                row[current_header["en"]] if current_header.get("en") is not None and current_header["en"] < len(row) else None,
            )
            cn = None
            if current_header.get("cn") is not None and current_header["cn"] < len(row):
                cn = row[current_header["cn"]]
            price = to_price(row[current_header["price"]]) if current_header.get("price") is not None and current_header["price"] < len(row) else None

            # Image lookup: same row or within +/- 1 row
            image_path = None
            for ar, ai in zip(anchor_rows, anchor_imgs):
                if abs(ar - row_idx_0) <= 1:
                    image_path = f"{IMG_DIR}/{ai}"
                    break

            if pn in parts:
                rec = parts[pn]
                # Keep first non-null values for descriptive fields.
                if not rec["english_name"] and en:
                    rec["english_name"] = str(en).strip()
                if not rec["chinese_name"] and cn:
                    rec["chinese_name"] = str(cn).strip()
                if rec["price"] is None and price is not None:
                    rec["price"] = price
                if not rec["image_path"] and image_path:
                    rec["image_path"] = image_path
                # Avoid duplicate fitment entries for same sheet
                if not any(f["sheet"] == fitment_entry["sheet"] for f in rec["fitment"]):
                    rec["fitment"].append(fitment_entry)
            else:
                parts[pn] = {
                    "part_number": pn,
                    "english_name": str(en).strip() if en else None,
                    "chinese_name": str(cn).strip() if cn else None,
                    "price": price,
                    "image_path": image_path,
                    "fitment": [fitment_entry],
                }
            added += 1
        sheet_part_counts[sname] = added
        print(f"  {sname!r}: {added} parts rows")

    wb.close()
    return parts, sheet_names, sheet_part_counts


# ----------------------------------------------------------------------------
# Step 3: write SQLite + JSON exports
# ----------------------------------------------------------------------------
def write_outputs(parts, sheet_names, sheet_part_counts):
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute("""
        CREATE TABLE parts (
            part_number  TEXT PRIMARY KEY,
            english_name TEXT,
            chinese_name TEXT,
            price        DECIMAL,
            image_path   TEXT,
            fitment      TEXT
        )
    """)
    for pn, rec in parts.items():
        cur.execute(
            "INSERT INTO parts (part_number, english_name, chinese_name, price, image_path, fitment) VALUES (?, ?, ?, ?, ?, ?)",
            (
                rec["part_number"],
                rec["english_name"],
                rec["chinese_name"],
                rec["price"],
                rec["image_path"],
                json.dumps(rec["fitment"], ensure_ascii=False),
            ),
        )
    con.commit()
    con.close()

    summary = {
        "total_parts": len(parts),
        "total_sheets": len(sheet_names),
        "sheets": [{"name": n, "parts_extracted": sheet_part_counts.get(n, 0)} for n in sheet_names],
        "images_extracted": len(os.listdir(IMG_DIR)),
    }
    with open(SUMMARY_PATH, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    sample = list(parts.values())[:SAMPLE_LIMIT]
    with open(SAMPLE_PATH, "w", encoding="utf-8") as f:
        json.dump(sample, f, ensure_ascii=False, indent=2)


def main():
    print(f"== Stage 1: extract images + drawing anchors ==")
    sheet_anchors = extract_images_and_anchors(XLSX)
    img_count = len(os.listdir(IMG_DIR))
    anchor_count = sum(len(v) for v in sheet_anchors.values())
    print(f"  images extracted: {img_count}")
    print(f"  total anchored images: {anchor_count}")

    print(f"\n== Stage 2: extract part rows ==")
    parts, sheet_names, counts = extract_parts(XLSX, sheet_anchors)

    print(f"\n== Stage 3: write outputs ==")
    write_outputs(parts, sheet_names, counts)
    print(f"  parts.db rows: {len(parts)}")
    print(f"  summary.json + sample_data.json written")
    print("\nDone.")


if __name__ == "__main__":
    main()
